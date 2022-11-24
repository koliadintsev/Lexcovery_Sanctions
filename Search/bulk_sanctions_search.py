import copy
import datetime
import os
import re

import openpyxl
from django.core.files.temp import NamedTemporaryFile
from openpyxl.styles import Font
from Lexcovery_Sanctions import settings
import requests
from dateutil.parser import parse
import aiohttp
import asyncio
from io import BytesIO
from Search import elasticsearch_handler, opencorporates_handler

SEARCH_LIST = os.path.join(settings.BASE_DIR,  'static') + "/Sanctions/Search/check.xlsx"
#SAVE_RESULT = os.path.join(settings.BASE_DIR,  'static') + "/Sanctions/Search/result.xlsx"
#SAVE_RESULT = os.path.join(settings.BASE_DIR,  'media') + "/Sanctions/Search/result.xlsx"
SAVE_RESULT = os.path.join(settings.BASE_DIR,  'tmp') + "/result.xlsx"
result_file = ''
entities_to_check=[]
check_results=[]


async def search_entities(file=SEARCH_LIST):
    global entities_to_check
    await import_entities_from_xls(file)
    await asyncio.gather(*[fuzzy_search(request) for request in entities_to_check])
    write_result()


async def import_entities_from_xls(file=SEARCH_LIST):
    global entities_to_check

    # Define variable to load the workbook
    workbook = openpyxl.load_workbook(file)

    # Define variable to read the active sheet:
    for sheet in workbook.worksheets:
        await asyncio.gather(*[import_from_entity_async(row) for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)])
    #print('import finished')


async def import_from_entity_async(row):
    global entities_to_check
    doc = []

    for value in row:
        if value is None:
            value = ''
        doc.append(value)

    result = {'id': doc[0], 'request': doc[1]}
    entities_to_check.append(result)


async def fuzzy_search(request):
    global check_results

    doc_id = request['id']
    entity = request['request']
    nominal = False

    result_auto = await elasticsearch_handler.search_fuzzy_request(entity, "AUTO")
    result_one = await elasticsearch_handler.search_fuzzy_request(entity, 2)

    result = result_auto
    for hit in result_one:
        found = False
        for old_hit in result_auto:
            if old_hit.main_name == hit.main_name and old_hit.sanctioned_by == hit.sanctioned_by:
                found = True
        if not found:
            result.append(hit)
    '''
    request_parts = entity.split()
    if len(request_parts) <=3:
        for s in request_parts:
            capital = re.sub('[^A-Z]', '', s)
            if len(capital) > 4:
                cap_result = await elasticsearch_handler.search_fuzzy_request(capital, 2)
                for hit in cap_result:
                    found = False
                    for old_hit in result:
                        if old_hit.main_name == hit.main_name and old_hit.sanctioned_by == hit.sanctioned_by:
                            found = True
                    if not found:
                        result.append(hit)
    '''
    try:
        officer_count = await opencorporates_handler.find_officer_count_by_name(entity)
    except Exception as e:
        officer_count = 0
    if officer_count > 5:
        nominal = True

    if len(result) == 0:
        return
    else:
        check = {'id': doc_id, 'request': entity, 'result': result, 'nominal': nominal}
        check_results.append(check)


def write_result():
    global check_results
    global result_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sanctions checking results"
    ws['A1'] = 'Request ID'
    ws['B1'] = 'Request'
    ws['C1'] = 'Hit Name'
    ws['D1'] = 'Sanctioned By'
    ws['E1'] = 'Program'
    ws['F1'] = 'Alternative Names'
    ws['G1'] = 'Details'
    ws['H1'] = 'Nationality'
    ws['I1'] = 'Address and Contacts'
    ws['J1'] = 'Additional Information'
    ws['K1'] = 'Nominal'
    header_row = ws.row_dimensions[1]
    header_row.font = Font(bold=True)

    last_row = 2
    for entity in check_results:
        for result in entity['result']:
            ws.cell(last_row, 1).value = entity['id']
            ws.cell(last_row, 2).value = entity['request']
            ws.cell(last_row, 3).value = result.main_name
            ws.cell(last_row, 4).value = result.sanctioned_by
            ws.cell(last_row, 5).value = result.program
            ws.cell(last_row, 6).value = result.names
            ws.cell(last_row, 7).value = result.personal_details
            ws.cell(last_row, 8).value = result.nationality
            ws.cell(last_row, 9).value = result.address
            ws.cell(last_row, 10).value = result.additional_info
            if entity['nominal']:
                ws.cell(last_row, 11).hyperlink = 'https://opencorporates.com/officers?jurisdiction_code=&q='+entity['request']+'&utf8=%E2%9C%93'
                ws.cell(last_row, 11).value = 'Possible nominal. Click for details.'
                ws.cell(last_row, 11).style = "Hyperlink"
            else:
                ws.cell(last_row, 11).value = 'No'
            last_row = last_row + 1

    #wb.save(SAVE_RESULT)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        result_file = tmp.read()


