import copy
import datetime
import os

import openpyxl
from openpyxl.styles import Font
from Lexcovery_Sanctions import settings
import requests
from dateutil.parser import parse
import aiohttp
import asyncio
from io import BytesIO
from Search import elasticsearch_handler
from DataImport import translit, import_companies_ua

SEARCH_LIST = os.path.join(settings.BASE_DIR,  'static') + "/Companies/Search/check.xlsx"
SAVE_RESULT = os.path.join(settings.BASE_DIR,  'static') + "/Companies/Search/result.xlsx"
entities_to_check=[]
check_results=[]
companies=[]


async def search_entities():
    global entities_to_check
    await import_entities_from_xls()
    await asyncio.gather(*[fuzzy_search_company(request) for request in entities_to_check])
    write_company()


async def search_founders():
    global entities_to_check
    global companies

    if len(companies) == 0:
        companies = await import_companies_ua.companies
        if len(companies) == 0:
            companies = await import_companies_ua.import_data_from_xml()

    await import_entities_from_xls()

    await asyncio.gather(*[search_founder(request) for request in entities_to_check])
    write_founder()


async def import_entities_from_xls():
    global entities_to_check
    entities_to_check = []

    # Define variable to load the workbook
    workbook = openpyxl.load_workbook(SEARCH_LIST)

    # Define variable to read the active sheet:
    for sheet in workbook.worksheets:
        await asyncio.gather(*[import_from_entity_async(row) for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)])
    #print('import finished')


async def import_from_entity_async(row):
    global entities_to_check
    doc = []

    for value in row:
        if value is None:
            value = ''
        if isinstance(value, str):
            if translit.is_latin(value):
                res = translit.to_cyrillic(value, 'ua')
                value = res
        doc.append(value)

    result = {'id': doc[0], 'request': doc[1]}
    entities_to_check.append(result)


async def fuzzy_search_company(request):
    global check_results

    doc_id = request['id']
    entity = request['request']

    check_old = {'id': doc_id, 'request': entity, 'result': []}
    check_results.append(check_old)
    index = check_results.index(check_old)

    result = elasticsearch_handler.find_ua_company(entity)
    if len(result) == 0:
        return
    else:
        check_new = {'id': doc_id, 'request': entity, 'result': result}
        check_results.remove(check_old)
        check_results.insert(index, check_new)


def write_company():
    global check_results

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Found companies"
    ws['A1'] = 'Request ID'
    ws['B1'] = 'Request'
    ws['C1'] = 'Hit Name'
    ws['D1'] = 'Hit Code'
    header_row = ws.row_dimensions[1]
    header_row.font = Font(bold=True)

    last_row = 2
    for entity in check_results:
        if len(entity['result']) == 0:
            ws.cell(last_row, 1).value = entity['id']
            ws.cell(last_row, 2).value = entity['request']
            last_row = last_row + 1
        else:
            for result in entity['result']:
                ws.cell(last_row, 1).value = entity['id']
                ws.cell(last_row, 2).value = entity['request']
                name = ''
                if result.shortname:
                    name = result.shortname
                else:
                    name = result.name
                ws.cell(last_row, 3).value = name
                ws.cell(last_row, 4).value = result.code
                last_row = last_row + 1

    wb.save(SAVE_RESULT)


async def search_founder(request):
    global check_results

    doc_id = request['id']
    entity = request['request']
    check = {'id': doc_id, 'result': entity}
    check_results.append(check)

    for company in companies:
        if company.name == entity or company.shortname == entity:

            await asyncio.gather(*[search_founder({'id': doc_id, 'request': founder.split(',')[0]}) for founder in
                                   company.founders])
            await asyncio.gather(*[search_founder({'id': doc_id, 'request': beneficiary.split(';')[0]}) for beneficiary in
                                   company.beneficiaries])


def write_founder():
    global check_results

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Founders and beneficiaries"
    ws['A1'] = 'Request ID'
    ws['B1'] = 'Hit Name'
    header_row = ws.row_dimensions[1]
    header_row.font = Font(bold=True)

    last_row = 2
    for entity in check_results:
        ws.cell(last_row, 1).value = entity['id']
        ws.cell(last_row, 2).value = entity['result']
        last_row = last_row + 1

    wb.save(SAVE_RESULT)
