import copy
import datetime
import os
import re

import openpyxl
import textract
from openpyxl.styles import Font
from Lexcovery_Sanctions import settings
import requests
from dateutil.parser import parse
import aiohttp
import asyncio
from io import BytesIO
from Search import elasticsearch_handler
from DataImport import translit, import_companies_ua
from bs4 import BeautifulSoup as bs

SEARCH_PATH = os.path.join(settings.BASE_DIR, 'static') + "/BG_Check/"


def get_files_list(path=SEARCH_PATH, format = ''):
    files = [os.path.join(path, f) for f in os.listdir(path) if
             (os.path.isfile(os.path.join(path, f)) and f.endswith(format))]

    return files


def parse_checks():
    companies = get_corp_structure()
    people = get_people_list()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items to check"
    ws['A1'] = 'Name'
    ws['B1'] = 'Shareholder'
    header_row = ws.row_dimensions[1]
    header_row.font = Font(bold=True)

    last_row = 2
    for entity in companies:
        for result in entity['shareholders']:
            ws.cell(last_row, 1).value = entity['entity']
            ws.cell(last_row, 2).value = result
            last_row = last_row + 1
    for entity in people:
        ws.cell(last_row, 1).value = entity['entity']
        ws.cell(last_row, 2).value = entity['shareholders']
        last_row = last_row + 1

    wb.save(SEARCH_PATH+'result.xlsx')


def get_corp_structure():
    files = get_files_list(path=SEARCH_PATH+'Companies/')
    companies = []
    for path in files:
        try:
            result = read_companies_file(path)
            companies.append(result)
        except Exception:
            print("ERROR PARSING: " + path)
    return companies


def get_people_list():
    files = get_files_list(path=SEARCH_PATH+'People/')
    people = []
    for path in files:
        try:
            result = read_people_file(path)
            people.append(result)
        except Exception:
            print("ERROR PARSING: " + path)
    return people


def read_companies_file(path=''):
    text = textract.process(path).decode()
    start = '3. Ultimate shareholders \(ultimate parent company\).?\n\n'
    end = '4.  Company’s  business  description'
    #end = '5. Key persons'
    #end = '\n\n'
    result = re.search('%s(.*?)%s' % (start, end), text, re.DOTALL)
    result = result.group(1)
    result = result.split('\n\n\n')[0]
    paragraphs = []
    results = result.split('\n')
    stop = -1
    max_spaces = -1
    max_tabs = -1

    for i in range(0, len(results)):
        if i == stop:
            continue
        p = results[i]
        s = p
        '''
        spaces = len(p) - len(p.lstrip())
        tabs = p.count('\t')
        if spaces < max_spaces and tabs <= max_tabs:
            break
        else:
            if spaces > max_spaces:
                max_spaces = spaces
            if tabs > max_tabs:
                max_tabs = tabs
        if i < len(results)-1:
            next_p = results[i+1]
            next_spaces = len(next_p) - len(p.lstrip())
            if spaces>next_spaces:
                s = s+next_p
                stop = i+1
        '''
        res = re.split('   | \t', s)
        for r in res:
            paragraphs.append(r)

    founders = []
    for s in paragraphs:
        if len(s)>0:
            s = re.sub('\((.*)\)', '', s)
            clear_string = re.sub("[–-−-] (.*)", '', s).strip()
            #clear_string = s.rsplit(' –')[0].rsplit(' -')[0].strip()
            if len(clear_string)>0:
                founders.append(clear_string)

    if len(founders) > 0:
        os.remove(path)
        return {'entity': founders[0], 'shareholders': founders}
    else:
        #if not result:
            #os.remove(path)
        raise Exception('ERROR PARSING')


def read_people_file(path=''):
    text = textract.process(path).decode()
    results = text.split('\n')
    result = ''
    for p in results:
        if p and not p == '\n':
            p = re.sub('\((.*)\)', '', p)
            p = re.sub('[,\.!?–-]', "", p)
            result = p.strip()
            break
    if result:
        os.remove(path)
        return {'entity': result, 'shareholders': result}
    else:
        print("ERROR PARSING: " + path)
        raise Exception('ERROR PARSING')

