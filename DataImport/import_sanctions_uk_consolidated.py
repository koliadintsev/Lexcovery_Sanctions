import copy
import datetime
import os

import openpyxl
from DataModel.UK import sanction_UK_consolidated
from Lexcovery_Sanctions import settings
from DataImport import translit
import requests
from dateutil.parser import parse
import aiohttp
import asyncio
from io import BytesIO

SANCTIONS_LIST = os.path.join(settings.BASE_DIR,  'static') + "/Sanctions/UK/ConList.xlsx"
sanctions = []
XLS_URL = "https://ofsistorage.blob.core.windows.net/publishlive/2022format/ConList.xlsx"
doc_id = 0


async def get_list_xls(session):
    #response = requests.get(XML_URL)
    response = await session.request(method='GET', url=XLS_URL)
    if response.ok:
        doc = await response.read()
        d = copy.deepcopy(doc)
        return d
    else:
        return


async def import_data_from_web(session):
    global sanctions

    # Define variable to load the workbook
    xls_file = await get_list_xls(session)
    workbook = openpyxl.load_workbook(BytesIO(xls_file))
    worksheet = workbook.active
    last_update = worksheet["B1"].value

    # Define variable to read the active sheet:
    await asyncio.gather(*[import_from_element_async(row) for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, values_only=True)])

    # print('import finished')

    return sanctions, last_update


async def import_from_element_async(row):
    global doc_id
    import_data_from_element(row, doc_id)
    doc_id = doc_id + 1


async def import_data_from_xls():
    global sanctions

    today = datetime.datetime.today()
    last_update = today.strftime("%d/%m/%Y")

    # Define variable to load the workbook
    workbook = openpyxl.load_workbook(SANCTIONS_LIST)
    worksheet = workbook.active
    last_update = worksheet["B1"]

    # Define variable to read the active sheet:
    await asyncio.gather(*[import_from_element_async(row) for row in
                           worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, values_only=True)])

    """
    tree = ET.fromstring(file.read().strip())
    executor = concurrent.futures.ThreadPoolExecutor(100)
    futures = [executor.submit(import_data_from_element, item, companies) for item in tree.findall('.//document')]
    concurrent.futures.wait(futures)
    """
    #print('import finished')
    return sanctions, last_update


def import_data_from_element(row, element_id):
    global sanctions
    doc = []

    for value in row:
        if value is None:
            value = ''
        doc.append(value)

    name = ''
    for i in range(1, 5):
        if doc[i]:
            name = name + doc[i] + ' '
    name = name + doc[0]
    title = doc[6]
    name_non_latin_script = doc[7]
    date_of_birth = doc[10]
    place_of_birth = doc[11] + ', ' + doc[12]
    nationality = doc[13]
    id_details = ''
    for i in range(14, 17):
        if doc[i]:
            id_details = id_details + ', '
    position = doc[18]
    address = ''
    for i in range(19, 26):
        if doc[i]:
            address = address + ', '
    additional_info = doc[27]
    group_id = doc[35]
    last_update = doc[34]
    designation_date = doc[33]
    listed_on = doc[32]
    regime = doc[31]
    alias_quality = doc[30]
    alias_type = doc[29]
    group_type = doc[28]

    sanction = sanction_UK_consolidated.SanctionUKConsolidated(name=name, title=title,
                                                               name_non_latin_script=name_non_latin_script,
                                                               date_of_birth=date_of_birth, place_of_birth=place_of_birth,
                                                               nationality=nationality, id_details=id_details, position=position,
                                                               address=address, additional_info=additional_info, group_type=group_type,
                                                               alias_type=alias_type, alias_quality=alias_quality, regime=regime,
                                                               listed_on=listed_on, designation_date=designation_date,
                                                               last_update=last_update, group_id=group_id, id=element_id)
    sanctions.append(sanction)


def import_data_from_json(element):

    element_id = element.id
    group_id = element.group_id
    last_update = element.last_update
    designation_date = element.designation_date
    listed_on = element.listed_on
    regime = element.regime
    alias_quality = element.alias_quality
    alias_type = element.alias_type
    group_type = element.group_type
    additional_info = element.additional_info
    address = element.address
    position = element.position
    id_details = element.id_details
    nationality = element.nationality
    place_of_birth = element.place_of_birth
    date_of_birth = element.date_of_birth
    name_non_latin_script = element.name_non_latin_script
    title = element.title
    name = element.name

    sanction = sanction_UK_consolidated.SanctionUKConsolidated(name=name, title=title,
                                                               name_non_latin_script=name_non_latin_script,
                                                               date_of_birth=date_of_birth, place_of_birth=place_of_birth,
                                                               nationality=nationality, id_details=id_details, position=position,
                                                               address=address, additional_info=additional_info, group_type=group_type,
                                                               alias_type=alias_type, alias_quality=alias_quality, regime=regime,
                                                               listed_on=listed_on, designation_date=designation_date,
                                                               last_update=last_update, group_id=group_id, id=element_id)
    return sanction

