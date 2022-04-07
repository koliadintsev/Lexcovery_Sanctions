import copy
import datetime
import os

import openpyxl
from DataModel.UA import sanction_ua
from Lexcovery_Sanctions import settings
from DataImport import translit
import requests
from dateutil.parser import parse
import aiohttp
import asyncio

SANCTIONS_LIST = os.path.join(settings.BASE_DIR,  'static') + "/Sanctions/UA/sanctions.xlsx"
sanctions = []
XLS_URL = "https://sanctions-t.rnbo.gov.ua/export/sanctions.xlsx"


async def get_list_xls(session):
    #response = requests.get(XML_URL)
    response = await session.request(method='GET', url=XLS_URL)
    if response.ok:
        doc = await response.read()
        d = copy.deepcopy(doc)
        return d
    else:
        return


async def import_data_from_web(session):
    global sanctions

    today = datetime.datetime.today()
    last_update = today.strftime("%d/%m/%Y")

    # Define variable to load the workbook
    workbook = openpyxl.load_workbook(SANCTIONS_LIST)

    # Define variable to read the active sheet:
    for sheet in workbook.worksheets:
        await asyncio.gather(*[import_from_element_async(row, sheet.title) for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)])

    # print('import finished')

    return sanctions, last_update


async def import_from_element_async(row, sheet_title):
    import_data_from_element(row, sheet_title)


async def import_data_from_xls():
    global sanctions

    today = datetime.datetime.today()
    last_update = today.strftime("%d/%m/%Y")

    # Define variable to load the workbook
    workbook = openpyxl.load_workbook(SANCTIONS_LIST)

    # Define variable to read the active sheet:
    for sheet in workbook.worksheets:
        await asyncio.gather(*[import_from_element_async(row, sheet.title) for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)])

    """
    tree = ET.fromstring(file.read().strip())
    executor = concurrent.futures.ThreadPoolExecutor(100)
    futures = [executor.submit(import_data_from_element, item, companies) for item in tree.findall('.//document')]
    concurrent.futures.wait(futures)
    """
    #print('import finished')
    return sanctions, last_update


def import_data_from_element(row, sheet_title):
    global sanctions
    doc = []

    for value in row:
        if value is None:
            value = ''
        doc.append(value)

    person = False
    if sheet_title == 'Фізичні особи':
        person = True

    id = doc[0]
    act_number = doc[1]
    start_date = doc[2]
    action = doc[3]
    changes = doc[4]
    number = doc[5]
    restrictions = doc[6]
    term = doc[7]
    end_date = doc[8]
    name_ukr = doc[9]
    name_orig = doc[10]
    name_alt = doc[11]
    name_latin = ''

    date_of_birth = ''
    citizenship = ''
    place_of_birth = ''
    work = ''
    responsive_body = ''
    iden_code = ''
    inn = ''
    address = ''
    address_additional = ''
    remarks = ''

    if translit.is_latin(name_orig):
        name_latin = name_orig
    elif translit.is_latin(name_alt):
        name_latin = name_alt
    else:
        if name_ukr:
            name_latin = translit.to_latin(name_ukr, 'ua')
        elif translit.is_cyrillic(name_orig):
            name_latin = translit.to_latin(name_orig, 'ru')
        elif translit.is_cyrillic(name_alt):
            name_latin = translit.to_latin(name_alt, 'ru')

    if person:
        date_of_birth = doc[12]
        citizenship = doc[13]
        place_of_birth = doc[14]
        work = doc[15]
        address = doc[16]
        remarks = doc[17]
        responsive_body = doc[18]
    else:
        iden_code = doc[12]
        inn = doc[13]
        address = doc[14]
        address_additional = doc[15]
        remarks = doc[16]
        responsive_body = doc[17]

    if not start_date:
        start_date = None
    if not end_date:
        end_date = None
    if not date_of_birth:
        date_of_birth = None

    sanction = sanction_ua.SanctionUA(act_number, start_date, action, changes, number, restrictions,
                 term, end_date, name_ukr, name_orig, name_alt, name_latin, date_of_birth,
                 citizenship, place_of_birth, work, address, address_additional, iden_code,
                 inn, remarks, responsive_body, id, person)
    sanctions.append(sanction)


def import_data_from_json(element):

    id = element.id
    act_number = element.act_number
    start_date = ''
    action = element.action
    changes = element.changes
    number = element.number
    restrictions = element.restrictions
    term = element.term
    end_date = ''
    name_ukr = element.name_ukr
    name_orig = element.name_orig
    name_alt = element.name_alt
    name_latin = element.name_latin
    date_of_birth = ''
    citizenship = element.citizenship
    place_of_birth = element.place_of_birth
    work = element.work
    responsive_body = element.responsive_body
    iden_code = element.iden_code
    inn = element.inn
    address = element.address
    address_additional = element.address_additional
    remarks = element.remarks
    person = element.person

    if element.start_date is not None:
        start_date = parse(element.start_date).date()
    if element.end_date is not None:
        end_date = parse(element.end_date).date()
    if element.date_of_birth is not None:
        date_of_birth = parse(element.date_of_birth).date()

    sanction = sanction_ua.SanctionUA(act_number, start_date, action, changes, number, restrictions,
                 term, end_date, name_ukr, name_orig, name_alt, name_latin, date_of_birth,
                 citizenship, place_of_birth, work, address, address_additional, iden_code,
                 inn, remarks, responsive_body, id, person)
    return sanction

